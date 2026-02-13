import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Yeni workbook oluştur
wb = openpyxl.Workbook()

# Varsayılan sheet'i kaldır
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ===== RENK VE STIL TANIMLARI =====
header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
title_font = Font(bold=True, size=14, color="1E293B")
currency_format = '#,##0.00 ₺'
border_style = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

# ===== 1. PERSONEL LİSTESİ SAYFASI =====
ws_personel = wb.create_sheet("Personel Listesi", 0)

# Başlık
ws_personel['A1'] = 'PERSONEL LİSTESİ'
ws_personel['A1'].font = title_font
ws_personel.merge_cells('A1:E1')
ws_personel['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_personel.row_dimensions[1].height = 30

# Tablo başlıkları
headers = ['Personel Adı Soyadı', 'Maaş', 'Yemek Yardımı', 'Yol Yardımı', 'Yıllık İzin Hakkı']
for col, header in enumerate(headers, 1):
    cell = ws_personel.cell(row=3, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_style

# Örnek veri satırları (10 satır boş)
for row in range(4, 14):
    for col in range(1, 6):
        cell = ws_personel.cell(row=row, column=col)
        cell.border = border_style
        if col in [2, 3, 4]:  # Para birimi sütunları
            cell.number_format = currency_format
        elif col == 5:  # İzin günü
            cell.number_format = '0 "Gün"'

# Sütun genişlikleri
ws_personel.column_dimensions['A'].width = 30
ws_personel.column_dimensions['B'].width = 15
ws_personel.column_dimensions['C'].width = 15
ws_personel.column_dimensions['D'].width = 15
ws_personel.column_dimensions['E'].width = 18

# ===== 2. PUANTAJ SAYFASI =====
ws_puantaj = wb.create_sheet("Puantaj", 1)

# Başlık
ws_puantaj['A1'] = 'PUANTAJ GİRİŞİ'
ws_puantaj['A1'].font = title_font
ws_puantaj.merge_cells('A1:J1')
ws_puantaj['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_puantaj.row_dimensions[1].height = 30

# Tablo başlıkları
puantaj_headers = ['Personel Adı', 'Maaş', 'Yemek', 'Yol', 'Prim', 'Mesai', 
                   'İzin Türü', 'İzin Gün', 'Kesinti', 'Net Hakediş']
for col, header in enumerate(puantaj_headers, 1):
    cell = ws_puantaj.cell(row=3, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_style

# Formül açıklamaları
ws_puantaj['L3'] = 'HESAPLAMA AÇIKLAMALARI:'
ws_puantaj['L3'].font = Font(bold=True, size=11)
ws_puantaj['L4'] = '• Aylık Çalışma Günü: 26 gün'
ws_puantaj['L5'] = '• Günlük Ücret = Maaş / 26'
ws_puantaj['L6'] = '• Saatlik Ücret = Günlük Ücret / 8'
ws_puantaj['L7'] = '• Yıllık İzinde kesinti YOK'
ws_puantaj['L8'] = '• Ücretsiz İzin/Rapor: Günlük ücretten düşer'
ws_puantaj['L9'] = '• Net Hakediş = Toplam Kazanç - Kesinti'

# Örnek veri satırları ve formüller
for row in range(4, 14):
    for col in range(1, 11):
        cell = ws_puantaj.cell(row=row, column=col)
        cell.border = border_style
        
        # Para birimi formatı
        if col in [2, 3, 4, 5, 6, 9, 10]:
            cell.number_format = currency_format
        
        # İzin gün formatı
        if col == 8:
            cell.number_format = '0'
    
    # Kesinti formülü (I sütunu)
    # Eğer izin türü "Ücretsiz" veya "Rapor" ise: (Maaş/26) * İzin Gün
    kesinti_cell = ws_puantaj.cell(row=row, column=9)
    kesinti_cell.value = f'=IF(OR(G{row}="Ücretsiz",G{row}="Rapor"),(B{row}/26)*H{row},0)'
    
    # Net Hakediş formülü (J sütunu)
    net_cell = ws_puantaj.cell(row=row, column=10)
    net_cell.value = f'=B{row}+C{row}+D{row}+E{row}+F{row}-I{row}'

# Sütun genişlikleri
ws_puantaj.column_dimensions['A'].width = 25
for col in ['B', 'C', 'D', 'E', 'F', 'I', 'J']:
    ws_puantaj.column_dimensions[col].width = 13
ws_puantaj.column_dimensions['G'].width = 12
ws_puantaj.column_dimensions['H'].width = 10

# ===== 3. BORDRO SAYFASI =====
ws_bordro = wb.create_sheet("Bordro", 2)

# Başlık bölümü
ws_bordro['A1'] = 'ÜCRET BORDROSU'
ws_bordro['A1'].font = Font(bold=True, size=16, color="1E293B")
ws_bordro.merge_cells('A1:D1')
ws_bordro['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_bordro.row_dimensions[1].height = 35

# Personel bilgileri
ws_bordro['A3'] = 'Personel Adı:'
ws_bordro['A3'].font = Font(bold=True)
ws_bordro['B3'] = '[Personel seçiniz]'

ws_bordro['A4'] = 'Dönem:'
ws_bordro['A4'].font = Font(bold=True)
ws_bordro['B4'] = '[Ay/Yıl]'

ws_bordro['A5'] = 'Tarih:'
ws_bordro['A5'].font = Font(bold=True)
ws_bordro['B5'] = datetime.now().strftime('%d.%m.%Y')

# KAZANÇLAR bölümü
ws_bordro['A7'] = 'KAZANÇLAR'
ws_bordro['A7'].font = Font(bold=True, size=12, color="FFFFFF")
ws_bordro['A7'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
ws_bordro.merge_cells('A7:B7')
ws_bordro['A7'].alignment = Alignment(horizontal='center')

kazanc_items = [
    ('Maaş', 'B8'),
    ('Yemek Yardımı', 'B9'),
    ('Yol Yardımı', 'B10'),
    ('Prim', 'B11'),
    ('Mesai', 'B12')
]

row = 8
for item, cell_ref in kazanc_items:
    ws_bordro[f'A{row}'] = item
    ws_bordro[f'A{row}'].border = border_style
    ws_bordro[cell_ref].number_format = currency_format
    ws_bordro[cell_ref].border = border_style
    row += 1

# Toplam Kazanç
ws_bordro['A13'] = 'TOPLAM KAZANÇ'
ws_bordro['A13'].font = Font(bold=True)
ws_bordro['A13'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
ws_bordro['B13'].value = '=SUM(B8:B12)'
ws_bordro['B13'].number_format = currency_format
ws_bordro['B13'].font = Font(bold=True)
ws_bordro['B13'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

# KESİNTİLER bölümü
ws_bordro['A15'] = 'KESİNTİLER'
ws_bordro['A15'].font = Font(bold=True, size=12, color="FFFFFF")
ws_bordro['A15'].fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
ws_bordro.merge_cells('A15:B15')
ws_bordro['A15'].alignment = Alignment(horizontal='center')

ws_bordro['A16'] = 'İzin Kesintisi'
ws_bordro['B16'].number_format = currency_format
ws_bordro['A16'].border = border_style
ws_bordro['B16'].border = border_style

ws_bordro['A17'] = 'Eksik Çalışma'
ws_bordro['B17'].number_format = currency_format
ws_bordro['A17'].border = border_style
ws_bordro['B17'].border = border_style

# Toplam Kesinti
ws_bordro['A18'] = 'TOPLAM KESİNTİ'
ws_bordro['A18'].font = Font(bold=True)
ws_bordro['A18'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
ws_bordro['B18'].value = '=SUM(B16:B17)'
ws_bordro['B18'].number_format = currency_format
ws_bordro['B18'].font = Font(bold=True)
ws_bordro['B18'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

# NET ÖDEME
ws_bordro['A20'] = 'NET ÖDEME'
ws_bordro['A20'].font = Font(bold=True, size=14, color="FFFFFF")
ws_bordro['A20'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
ws_bordro['A20'].alignment = Alignment(horizontal='center', vertical='center')
ws_bordro['B20'].value = '=B13-B18'
ws_bordro['B20'].number_format = currency_format
ws_bordro['B20'].font = Font(bold=True, size=14, color="10B981")
ws_bordro['B20'].fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
ws_bordro['B20'].alignment = Alignment(horizontal='right', vertical='center')
ws_bordro.row_dimensions[20].height = 30

# Sütun genişlikleri
ws_bordro.column_dimensions['A'].width = 25
ws_bordro.column_dimensions['B'].width = 20

# İmza bölümü
ws_bordro['A23'] = 'Personel İmzası: _______________'
ws_bordro['A23'].alignment = Alignment(horizontal='left')

ws_bordro['A25'] = 'Yetkili İmzası: _______________'
ws_bordro['A25'].alignment = Alignment(horizontal='left')

# Dosyayı kaydet
output_file = r'c:\Users\burha\Desktop\Yeni Uygumalar\personeltakip\Personel_Takip_Bordro.xlsx'
wb.save(output_file)
print(f"Excel dosyası başarıyla oluşturuldu: {output_file}")
